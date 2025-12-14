import os
import re
import difflib
from collections import defaultdict
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import datetime


def normalize_triplet(triplet_str):
    """
    标准化三元组字符串，移除括号和多余空格，统一格式
    """
    # 移除所有类型的括号
    normalized = re.sub(r'[\(\)\[\]\{\}]', '', triplet_str)
    # 移除多余空格并标准化逗号分隔符
    normalized = re.sub(r'\s*,\s*', ', ', normalized)
    # 统一实体格式（如去除括号内的注释）
    normalized = re.sub(r'\([^)]*\)', '', normalized)  # 移除括号内的内容
    normalized = re.sub(r'\s+', ' ', normalized)  # 合并多个空格
    return normalized.strip()


def parse_triplets(content):
    """
    从文本内容中解析三元组
    返回三元组列表，每个三元组格式为(主体, 关系, 客体)
    """
    triplets = []

    # 使用正则表达式匹配三元组，允许不同类型的括号
    pattern = r'[\(\[\{]([^\)\]\}]+?)[\)\]\}]'
    matches = re.findall(pattern, content)

    for match in matches:
        # 分割三元组为三个部分
        parts = [part.strip() for part in match.split(',')]
        if len(parts) >= 3:
            subject = parts[0]
            predicate = parts[1]
            obj = ', '.join(parts[2:])  # 处理客体中可能包含的逗号
            triplets.append((subject, predicate, obj))

    return triplets


def extract_entity_type(entity):
    """
    从实体字符串中提取实体类型
    例如: "中成药:龙胆泻肝丸" -> "中成药"
    """
    if ':' in entity:
        return entity.split(':', 1)[0]
    return "未知类型"


def extract_relation_type(relation):
    """
    提取关系类型（直接返回关系字符串）
    """
    return relation


def is_similar_entity(entity1, entity2, threshold=0.8):
    """
    判断两个实体是否相似
    使用字符串相似度算法，阈值可调整
    """
    # 如果完全相同，直接返回True
    if entity1 == entity2:
        return True

    # 提取实体内容（去除类型前缀）
    content1 = entity1.split(':', 1)[1] if ':' in entity1 else entity1
    content2 = entity2.split(':', 1)[1] if ':' in entity2 else entity2

    # 使用difflib计算相似度
    similarity = difflib.SequenceMatcher(None, content1, content2).ratio()

    # 检查是否一个实体包含另一个实体
    if content1 in content2 or content2 in content1:
        return True

    # 检查相似度是否超过阈值
    return similarity >= threshold


def deduplicate_entities(entities, threshold=0.8):
    """
    对实体列表进行去重，基于相似性判断
    """
    unique_entities = []

    for entity in entities:
        # 检查是否已经存在相似的实体
        found_similar = False
        for unique_entity in unique_entities:
            if is_similar_entity(entity, unique_entity, threshold):
                found_similar = True
                break

        if not found_similar:
            unique_entities.append(entity)

    return unique_entities


def calculate_metrics(gold_triplets, pred_triplets):
    """
    计算精确度、召回率和F1分数，并返回未匹配的三元组
    """
    # 标准化三元组
    gold_normalized = [normalize_triplet(f"({s}, {p}, {o})") for s, p, o in gold_triplets]
    pred_normalized = [normalize_triplet(f"({s}, {p}, {o})") for s, p, o in pred_triplets]

    # 计算TP, FP, FN（不进行查重）
    tp = 0  # 正确识别的三元组
    fp = 0  # 错误识别的三元组
    fn = 0  # 未识别的三元组

    # 存储未匹配的三元组
    unmatched_gold_triplets = []  # FN: 标准中有但预测中没有的
    unmatched_pred_triplets = []  # FP: 预测中有但标准中没有的

    # 创建预测三元组的副本用于匹配
    pred_remaining = pred_normalized.copy()

    # 计算TP和FN
    for i, gold_triplet in enumerate(gold_normalized):
        matched = False
        for j, pred_triplet in enumerate(pred_remaining):
            if gold_triplet == pred_triplet:
                tp += 1
                pred_remaining.pop(j)
                matched = True
                break
        if not matched:
            fn += 1
            unmatched_gold_triplets.append(gold_triplets[i])  # 保存原始三元组

    # 剩余的都是FP
    fp = len(pred_remaining)
    # 保存FP的三元组
    for pred_triplet in pred_remaining:
        # 找到原始预测三元组
        for i, triplet in enumerate(pred_triplets):
            if normalize_triplet(f"({triplet[0]}, {triplet[1]}, {triplet[2]})") == pred_triplet:
                unmatched_pred_triplets.append(triplet)
                break

    # 计算指标
    precision = tp / (tp + fp) if (tp + fp) > 0 else 0
    recall = tp / (tp + fn) if (tp + fn) > 0 else 0
    f1 = 2 * precision * recall / (precision + recall) if (precision + recall) > 0 else 0

    return {
        'tp': tp,
        'fp': fp,
        'fn': fn,
        'precision': precision,
        'recall': recall,
        'f1': f1,
        'unmatched_gold': unmatched_gold_triplets,
        'unmatched_pred': unmatched_pred_triplets
    }


def extract_entities_and_relations(triplets, deduplicate_entities_flag=True):
    """
    从三元组列表中提取所有实体和关系
    实体可以选择是否去重，关系不去重
    """
    entities = []
    relations = []

    for s, p, o in triplets:
        entities.append(s)
        entities.append(o)
        relations.append(p)

    # 实体去重
    if deduplicate_entities_flag:
        entities = deduplicate_entities(entities)

    return entities, relations


def calculate_entity_metrics(gold_entities, pred_entities):
    """
    计算实体级别的指标（已去重），并返回未匹配的实体
    支持相似实体匹配
    """
    # 计算TP, FP, FN
    tp = 0  # 正确识别的实体
    fp = 0  # 错误识别的实体
    fn = 0  # 未识别的实体

    # 存储未匹配的实体
    unmatched_gold_entities = []  # FN: 标准中有但预测中没有的
    unmatched_pred_entities = []  # FP: 预测中有但标准中没有的

    # 按实体类型分类统计
    entity_type_stats = defaultdict(lambda: {'tp': 0, 'fp': 0, 'fn': 0})

    # 创建预测实体的副本用于匹配
    pred_remaining = pred_entities.copy()

    # 计算TP和FN
    for gold_entity in gold_entities:
        matched = False
        entity_type = extract_entity_type(gold_entity)

        for i, pred_entity in enumerate(pred_remaining):
            if is_similar_entity(gold_entity, pred_entity):
                tp += 1
                pred_remaining.pop(i)
                matched = True
                entity_type_stats[entity_type]['tp'] += 1
                break
        if not matched:
            fn += 1
            unmatched_gold_entities.append(gold_entity)
            entity_type_stats[entity_type]['fn'] += 1

    # 剩余的都是FP
    fp = len(pred_remaining)
    unmatched_pred_entities = pred_remaining

    # 统计FP的实体类型
    for pred_entity in pred_remaining:
        entity_type = extract_entity_type(pred_entity)
        entity_type_stats[entity_type]['fp'] += 1

    # 计算每种实体类型的指标
    for entity_type, stats in entity_type_stats.items():
        tp_type = stats['tp']
        fp_type = stats['fp']
        fn_type = stats['fn']

        precision_type = tp_type / (tp_type + fp_type) if (tp_type + fp_type) > 0 else 0
        recall_type = tp_type / (tp_type + fn_type) if (tp_type + fn_type) > 0 else 0
        f1_type = 2 * precision_type * recall_type / (precision_type + recall_type) if (
                                                                                               precision_type + recall_type) > 0 else 0

        stats['precision'] = precision_type
        stats['recall'] = recall_type
        stats['f1'] = f1_type

    # 计算总体指标
    precision = tp / (tp + fp) if (tp + fp) > 0 else 0
    recall = tp / (tp + fn) if (tp + fn) > 0 else 0
    f1 = 2 * precision * recall / (precision + recall) if (precision + recall) > 0 else 0

    return {
        'tp': tp,
        'fp': fp,
        'fn': fn,
        'precision': precision,
        'recall': recall,
        'f1': f1,
        'unmatched_gold': unmatched_gold_entities,
        'unmatched_pred': unmatched_pred_entities,
        'type_stats': entity_type_stats
    }


def calculate_relation_metrics(gold_relations, pred_relations):
    """
    计算关系级别的指标（不进行查重），并返回未匹配的关系
    """
    # 计算TP, FP, FN
    tp = 0  # 正确识别的关系
    fp = 0  # 错误识别的关系
    fn = 0  # 未识别的关系

    # 存储未匹配的关系
    unmatched_gold_relations = []  # FN: 标准中有但预测中没有的
    unmatched_pred_relations = []  # FP: 预测中有但标准中没有的

    # 按关系类型分类统计
    relation_type_stats = defaultdict(lambda: {'tp': 0, 'fp': 0, 'fn': 0})

    # 创建预测关系的副本用于匹配
    pred_remaining = pred_relations.copy()

    # 计算TP和FN
    for gold_relation in gold_relations:
        matched = False
        relation_type = extract_relation_type(gold_relation)

        for i, pred_relation in enumerate(pred_remaining):
            if gold_relation == pred_relation:
                tp += 1
                pred_remaining.pop(i)
                matched = True
                relation_type_stats[relation_type]['tp'] += 1
                break
        if not matched:
            fn += 1
            unmatched_gold_relations.append(gold_relation)
            relation_type_stats[relation_type]['fn'] += 1

    # 剩余的都是FP
    fp = len(pred_remaining)
    unmatched_pred_relations = pred_remaining

    # 统计FP的关系类型
    for pred_relation in pred_remaining:
        relation_type = extract_relation_type(pred_relation)
        relation_type_stats[relation_type]['fp'] += 1

    # 计算每种关系类型的指标
    for relation_type, stats in relation_type_stats.items():
        tp_type = stats['tp']
        fp_type = stats['fp']
        fn_type = stats['fn']

        precision_type = tp_type / (tp_type + fp_type) if (tp_type + fp_type) > 0 else 0
        recall_type = tp_type / (tp_type + fn_type) if (tp_type + fn_type) > 0 else 0
        f1_type = 2 * precision_type * recall_type / (precision_type + recall_type) if (
                                                                                               precision_type + recall_type) > 0 else 0

        stats['precision'] = precision_type
        stats['recall'] = recall_type
        stats['f1'] = f1_type

    # 计算总体指标
    precision = tp / (tp + fp) if (tp + fp) > 0 else 0
    recall = tp / (tp + fn) if (tp + fn) > 0 else 0
    f1 = 2 * precision * recall / (precision + recall) if (precision + recall) > 0 else 0

    return {
        'tp': tp,
        'fp': fp,
        'fn': fn,
        'precision': precision,
        'recall': recall,
        'f1': f1,
        'unmatched_gold': unmatched_gold_relations,
        'unmatched_pred': unmatched_pred_relations,
        'type_stats': relation_type_stats
    }


def process_file_pair(gold_file_path, pred_file_path):
    """
    处理一对文件（人工标注和模型输出）
    """
    try:
        # 读取文件内容
        with open(gold_file_path, 'r', encoding='utf-8') as f:
            gold_content = f.read()

        with open(pred_file_path, 'r', encoding='utf-8') as f:
            pred_content = f.read()

        # 解析三元组
        gold_triplets = parse_triplets(gold_content)
        pred_triplets = parse_triplets(pred_content)

        # 计算三元组级别的指标
        triplet_metrics = calculate_metrics(gold_triplets, pred_triplets)

        # 提取实体和关系（实体需要去重，关系不需要）
        gold_entities, gold_relations = extract_entities_and_relations(gold_triplets, deduplicate_entities_flag=True)
        pred_entities, pred_relations = extract_entities_and_relations(pred_triplets, deduplicate_entities_flag=True)

        # 计算实体指标
        entity_metrics = calculate_entity_metrics(gold_entities, pred_entities)

        # 计算关系指标
        relation_metrics = calculate_relation_metrics(gold_relations, pred_relations)

        # 返回所有指标
        return {
            'gold_triplets': gold_triplets,
            'pred_triplets': pred_triplets,
            'gold_entities': gold_entities,
            'pred_entities': pred_entities,
            'gold_relations': gold_relations,
            'pred_relations': pred_relations,
            'triplet': triplet_metrics,
            'entity': entity_metrics,
            'relation': relation_metrics
        }

    except Exception as e:
        print(f"处理文件时出错: {e}")
        return None


def save_results_to_excel(results, output_file="评估结果.xlsx"):
    """
    将评估结果保存到Excel文件
    """
    try:
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # 1. 创建汇总工作表
            summary_data = []
            for file_name, file_results in results['file_results'].items():
                summary_data.append({
                    '文件名': file_name,
                    '标准三元组数': len(file_results['gold_triplets']),
                    '预测三元组数': len(file_results['pred_triplets']),
                    '三元组TP': file_results['triplet']['tp'],
                    '三元组FP': file_results['triplet']['fp'],
                    '三元组FN': file_results['triplet']['fn'],
                    '三元组精确度': file_results['triplet']['precision'],
                    '三元组召回率': file_results['triplet']['recall'],
                    '三元组F1': file_results['triplet']['f1'],
                    '实体TP': file_results['entity']['tp'],
                    '实体FP': file_results['entity']['fp'],
                    '实体FN': file_results['entity']['fn'],
                    '实体精确度': file_results['entity']['precision'],
                    '实体召回率': file_results['entity']['recall'],
                    '实体F1': file_results['entity']['f1'],
                    '关系TP': file_results['relation']['tp'],
                    '关系FP': file_results['relation']['fp'],
                    '关系FN': file_results['relation']['fn'],
                    '关系精确度': file_results['relation']['precision'],
                    '关系召回率': file_results['relation']['recall'],
                    '关系F1': file_results['relation']['f1']
                })

            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name='文件汇总', index=False)

            # 2. 创建总体指标工作表
            overall_data = []
            # 三元组总体指标
            overall_data.append({
                '指标类型': '三元组',
                'TP': results['overall_triplet']['tp'],
                'FP': results['overall_triplet']['fp'],
                'FN': results['overall_triplet']['fn'],
                '精确度': results['overall_triplet']['precision'],
                '召回率': results['overall_triplet']['recall'],
                'F1分数': results['overall_triplet']['f1']
            })
            # 实体总体指标
            overall_data.append({
                '指标类型': '实体',
                'TP': results['overall_entity']['tp'],
                'FP': results['overall_entity']['fp'],
                'FN': results['overall_entity']['fn'],
                '精确度': results['overall_entity']['precision'],
                '召回率': results['overall_entity']['recall'],
                'F1分数': results['overall_entity']['f1']
            })
            # 关系总体指标
            overall_data.append({
                '指标类型': '关系',
                'TP': results['overall_relation']['tp'],
                'FP': results['overall_relation']['fp'],
                'FN': results['overall_relation']['fn'],
                '精确度': results['overall_relation']['precision'],
                '召回率': results['overall_relation']['recall'],
                'F1分数': results['overall_relation']['f1']
            })

            overall_df = pd.DataFrame(overall_data)
            overall_df.to_excel(writer, sheet_name='总体指标', index=False)

            # 3. 创建实体类型指标工作表
            entity_type_data = []
            for entity_type, stats in results['overall_entity_type_stats'].items():
                entity_type_data.append({
                    '实体类型': entity_type,
                    'TP': stats['tp'],
                    'FP': stats['fp'],
                    'FN': stats['fn'],
                    '精确度': stats['precision'],
                    '召回率': stats['recall'],
                    'F1分数': stats['f1']
                })

            entity_type_df = pd.DataFrame(entity_type_data)
            entity_type_df.to_excel(writer, sheet_name='实体类型指标', index=False)

            # 4. 创建关系类型指标工作表
            relation_type_data = []
            for relation_type, stats in results['overall_relation_type_stats'].items():
                relation_type_data.append({
                    '关系类型': relation_type,
                    'TP': stats['tp'],
                    'FP': stats['fp'],
                    'FN': stats['fn'],
                    '精确度': stats['precision'],
                    '召回率': stats['recall'],
                    'F1分数': stats['f1']
                })

            relation_type_df = pd.DataFrame(relation_type_data)
            relation_type_df.to_excel(writer, sheet_name='关系类型指标', index=False)

            # 5. 创建中成药特别分析工作表
            cm_analysis_data = []
            # 总体中成药指标
            if "中成药" in results['overall_entity_type_stats']:
                cm_stats = results['overall_entity_type_stats']["中成药"]
                cm_analysis_data.append({
                    '指标类型': '总体指标',
                    'TP': cm_stats['tp'],
                    'FP': cm_stats['fp'],
                    'FN': cm_stats['fn'],
                    '精确度': cm_stats['precision'],
                    '召回率': cm_stats['recall'],
                    'F1分数': cm_stats['f1']
                })

            cm_analysis_df = pd.DataFrame(cm_analysis_data)
            cm_analysis_df.to_excel(writer, sheet_name='中成药分析', index=False)

        print(f"\n评估结果已保存到: {output_file}")

        # 设置Excel格式
        wb = writer.book
        # 设置总体指标的格式
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            # 设置列宽
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            # 设置标题行样式
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            header_alignment = Alignment(horizontal="center", vertical="center")

            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment

        # 保存文件
        wb.save(output_file)

    except Exception as e:
        print(f"保存Excel文件时出错: {e}")


def process_folders(gold_folder, pred_folder, save_excel=True, output_file=None):
    """
    处理两个文件夹中的所有txt文件
    """
    # 获取文件夹中的所有txt文件
    gold_files = [f for f in os.listdir(gold_folder) if f.endswith('.txt')]
    pred_files = [f for f in os.listdir(pred_folder) if f.endswith('.txt')]

    # 找出两个文件夹中都存在的文件
    common_files = set(gold_files) & set(pred_files)

    if not common_files:
        print("没有找到名称相同的txt文件进行匹配")
        return None

    # 初始化总体统计
    overall_triplet_metrics = {'tp': 0, 'fp': 0, 'fn': 0}
    overall_entity_metrics = {'tp': 0, 'fp': 0, 'fn': 0}
    overall_relation_metrics = {'tp': 0, 'fp': 0, 'fn': 0}

    # 初始化类型级别统计
    overall_entity_type_stats = defaultdict(lambda: {'tp': 0, 'fp': 0, 'fn': 0})
    overall_relation_type_stats = defaultdict(lambda: {'tp': 0, 'fp': 0, 'fn': 0})

    # 初始化中成药未匹配实体统计
    overall_cm_fn_entities = []  # 中成药FN实体（标准中有但预测中没有）
    overall_cm_fp_entities = []  # 中成药FP实体（预测中有但标准中没有）

    # 存储每个文件的详细结果
    file_results = {}

    # 处理每个文件
    for file_name in sorted(common_files):
        print("\n" + "=" * 60)
        print(f"处理文件: {file_name}")
        print("=" * 60)

        gold_file_path = os.path.join(gold_folder, file_name)
        pred_file_path = os.path.join(pred_folder, file_name)

        # 处理文件对
        metrics = process_file_pair(gold_file_path, pred_file_path)

        if metrics:
            # 打印结果
            print(f"人工标注三元组数量: {len(metrics['gold_triplets'])}")
            print(f"模型预测三元组数量: {len(metrics['pred_triplets'])}")
            print(f"去重后标准实体数量: {len(metrics['gold_entities'])}")
            print(f"去重后预测实体数量: {len(metrics['pred_entities'])}")
            print(f"标准关系数量: {len(metrics['gold_relations'])}")
            print(f"预测关系数量: {len(metrics['pred_relations'])}")

            print("\n三元组识别结果:")
            print(f"  TP: {metrics['triplet']['tp']}, FP: {metrics['triplet']['fp']}, FN: {metrics['triplet']['fn']}")
            print(
                f"  精确度: {metrics['triplet']['precision']:.4f}, 召回率: {metrics['triplet']['recall']:.4f}, F1分数: {metrics['triplet']['f1']:.4f}")

            print("\n实体识别结果:")
            print(f"  TP: {metrics['entity']['tp']}, FP: {metrics['entity']['fp']}, FN: {metrics['entity']['fn']}")
            print(
                f"  精确度: {metrics['entity']['precision']:.4f}, 召回率: {metrics['entity']['recall']:.4f}, F1分数: {metrics['entity']['f1']:.4f}")

            print("\n关系识别结果:")
            print(
                f"  TP: {metrics['relation']['tp']}, FP: {metrics['relation']['fp']}, FN: {metrics['relation']['fn']}")
            print(
                f"  精确度: {metrics['relation']['precision']:.4f}, 召回率: {metrics['relation']['recall']:.4f}, F1分数: {metrics['relation']['f1']:.4f}")

            # 打印每种实体类型的指标
            print("\n每种实体类型的指标:")
            for entity_type, stats in metrics['entity']['type_stats'].items():
                print(f"  {entity_type}: TP={stats['tp']}, FP={stats['fp']}, FN={stats['fn']}, "
                      f"精确度={stats['precision']:.4f}, 召回率={stats['recall']:.4f}, F1={stats['f1']:.4f}")

            # 打印每种关系类型的指标
            print("\n每种关系类型的指标:")
            for relation_type, stats in metrics['relation']['type_stats'].items():
                print(f"  {relation_type}: TP={stats['tp']}, FP={stats['fp']}, FN={stats['fn']}, "
                      f"精确度={stats['precision']:.4f}, 召回率={stats['recall']:.4f}, F1={stats['f1']:.4f}")

            # 打印未匹配的三元组
            print("\n未匹配的三元组 (FN - 标准中有但预测中没有):")
            for triplet in metrics['triplet']['unmatched_gold']:
                print(f"  ({triplet[0]}, {triplet[1]}, {triplet[2]})")

            print("\n未匹配的三元组 (FP - 预测中有但标准中没有):")
            for triplet in metrics['triplet']['unmatched_pred']:
                print(f"  ({triplet[0]}, {triplet[1]}, {triplet[2]})")

            # 打印未匹配的实体
            print("\n未匹配的实体 (FN - 标准中有但预测中没有):")
            for entity in metrics['entity']['unmatched_gold']:
                print(f"  {entity}")

            print("\n未匹配的实体 (FP - 预测中有但标准中没有):")
            for entity in metrics['entity']['unmatched_pred']:
                print(f"  {entity}")

            # 打印未匹配的关系
            print("\n未匹配的关系 (FN - 标准中有但预测中没有):")
            for relation in metrics['relation']['unmatched_gold']:
                print(f"  {relation}")

            print("\n未匹配的关系 (FP - 预测中有但标准中没有):")
            for relation in metrics['relation']['unmatched_pred']:
                print(f"  {relation}")

            # 存储文件结果
            file_results[file_name] = metrics

            # 累加三元组统计信息（三元组不需要去重）
            overall_triplet_metrics['tp'] += metrics['triplet']['tp']
            overall_triplet_metrics['fp'] += metrics['triplet']['fp']
            overall_triplet_metrics['fn'] += metrics['triplet']['fn']

            # 累加实体统计信息（每个文件内已去重）
            overall_entity_metrics['tp'] += metrics['entity']['tp']
            overall_entity_metrics['fp'] += metrics['entity']['fp']
            overall_entity_metrics['fn'] += metrics['entity']['fn']

            # 累加关系统计信息（关系不需要去重）
            overall_relation_metrics['tp'] += metrics['relation']['tp']
            overall_relation_metrics['fp'] += metrics['relation']['fp']
            overall_relation_metrics['fn'] += metrics['relation']['fn']

            # 累加实体类型统计（每个文件内已去重）
            for entity_type, stats in metrics['entity']['type_stats'].items():
                overall_entity_type_stats[entity_type]['tp'] += stats['tp']
                overall_entity_type_stats[entity_type]['fp'] += stats['fp']
                overall_entity_type_stats[entity_type]['fn'] += stats['fn']

            # 累加关系类型统计
            for relation_type, stats in metrics['relation']['type_stats'].items():
                overall_relation_type_stats[relation_type]['tp'] += stats['tp']
                overall_relation_type_stats[relation_type]['fp'] += stats['fp']
                overall_relation_type_stats[relation_type]['fn'] += stats['fn']

            # 收集该文件的中成药未匹配实体
            entity_metrics = metrics['entity']
            # 从该文件的实体未匹配中提取中成药实体
            for entity in entity_metrics['unmatched_gold']:
                if entity.startswith('中成药:'):
                    overall_cm_fn_entities.append(entity)

            for entity in entity_metrics['unmatched_pred']:
                if entity.startswith('中成药:'):
                    overall_cm_fp_entities.append(entity)

    # 对中成药未匹配实体进行去重
    overall_cm_fn_entities = deduplicate_entities(overall_cm_fn_entities)
    overall_cm_fp_entities = deduplicate_entities(overall_cm_fp_entities)

    # 计算总体指标
    triplet_precision = overall_triplet_metrics['tp'] / (
            overall_triplet_metrics['tp'] + overall_triplet_metrics['fp']) if (overall_triplet_metrics['tp'] +
                                                                               overall_triplet_metrics[
                                                                                   'fp']) > 0 else 0
    triplet_recall = overall_triplet_metrics['tp'] / (
            overall_triplet_metrics['tp'] + overall_triplet_metrics['fn']) if (overall_triplet_metrics['tp'] +
                                                                               overall_triplet_metrics[
                                                                                   'fn']) > 0 else 0
    triplet_f1 = 2 * triplet_precision * triplet_recall / (triplet_precision + triplet_recall) if (
                                                                                                          triplet_precision + triplet_recall) > 0 else 0

    entity_precision = overall_entity_metrics['tp'] / (overall_entity_metrics['tp'] + overall_entity_metrics['fp']) if (
                                                                                                                               overall_entity_metrics[
                                                                                                                                   'tp'] +
                                                                                                                               overall_entity_metrics[
                                                                                                                                   'fp']) > 0 else 0
    entity_recall = overall_entity_metrics['tp'] / (overall_entity_metrics['tp'] + overall_entity_metrics['fn']) if (
                                                                                                                            overall_entity_metrics[
                                                                                                                                'tp'] +
                                                                                                                            overall_entity_metrics[
                                                                                                                                'fn']) > 0 else 0
    entity_f1 = 2 * entity_precision * entity_recall / (entity_precision + entity_recall) if (
                                                                                                     entity_precision + entity_recall) > 0 else 0

    relation_precision = overall_relation_metrics['tp'] / (
            overall_relation_metrics['tp'] + overall_relation_metrics['fp']) if (overall_relation_metrics['tp'] +
                                                                                 overall_relation_metrics[
                                                                                     'fp']) > 0 else 0
    relation_recall = overall_relation_metrics['tp'] / (
            overall_relation_metrics['tp'] + overall_relation_metrics['fn']) if (overall_relation_metrics['tp'] +
                                                                                 overall_relation_metrics[
                                                                                     'fn']) > 0 else 0
    relation_f1 = 2 * relation_precision * relation_recall / (relation_precision + relation_recall) if (
                                                                                                               relation_precision + relation_recall) > 0 else 0

    # 计算每种实体类型的总体指标
    for entity_type, stats in overall_entity_type_stats.items():
        tp_type = stats['tp']
        fp_type = stats['fp']
        fn_type = stats['fn']

        precision_type = tp_type / (tp_type + fp_type) if (tp_type + fp_type) > 0 else 0
        recall_type = tp_type / (tp_type + fn_type) if (tp_type + fn_type) > 0 else 0
        f1_type = 2 * precision_type * recall_type / (precision_type + recall_type) if (
                                                                                               precision_type + recall_type) > 0 else 0

        stats['precision'] = precision_type
        stats['recall'] = recall_type
        stats['f1'] = f1_type

    # 计算每种关系类型的总体指标
    for relation_type, stats in overall_relation_type_stats.items():
        tp_type = stats['tp']
        fp_type = stats['fp']
        fn_type = stats['fn']

        precision_type = tp_type / (tp_type + fp_type) if (tp_type + fp_type) > 0 else 0
        recall_type = tp_type / (tp_type + fn_type) if (tp_type + fn_type) > 0 else 0
        f1_type = 2 * precision_type * recall_type / (precision_type + recall_type) if (
                                                                                               precision_type + recall_type) > 0 else 0

        stats['precision'] = precision_type
        stats['recall'] = recall_type
        stats['f1'] = f1_type

    # 准备总体指标字典
    overall_triplet = {
        'tp': overall_triplet_metrics['tp'],
        'fp': overall_triplet_metrics['fp'],
        'fn': overall_triplet_metrics['fn'],
        'precision': triplet_precision,
        'recall': triplet_recall,
        'f1': triplet_f1
    }

    overall_entity = {
        'tp': overall_entity_metrics['tp'],
        'fp': overall_entity_metrics['fp'],
        'fn': overall_entity_metrics['fn'],
        'precision': entity_precision,
        'recall': entity_recall,
        'f1': entity_f1
    }

    overall_relation = {
        'tp': overall_relation_metrics['tp'],
        'fp': overall_relation_metrics['fp'],
        'fn': overall_relation_metrics['fn'],
        'precision': relation_precision,
        'recall': relation_recall,
        'f1': relation_f1
    }

    # 打印总体结果
    print("\n" + "=" * 60)
    print("总体性能评估结果")
    print("=" * 60)

    print("\n三元组识别总体结果:")
    print(f"  TP: {overall_triplet['tp']}, FP: {overall_triplet['fp']}, FN: {overall_triplet['fn']}")
    print(f"  精确度: {overall_triplet['precision']:.4f}")
    print(f"  召回率: {overall_triplet['recall']:.4f}")
    print(f"  F1分数: {overall_triplet['f1']:.4f}")

    print("\n实体识别总体结果:")
    print(f"  TP: {overall_entity['tp']}, FP: {overall_entity['fp']}, FN: {overall_entity['fn']}")
    print(f"  精确度: {overall_entity['precision']:.4f}")
    print(f"  召回率: {overall_entity['recall']:.4f}")
    print(f"  F1分数: {overall_entity['f1']:.4f}")

    print("\n关系识别总体结果:")
    print(f"  TP: {overall_relation['tp']}, FP: {overall_relation['fp']}, FN: {overall_relation['fn']}")
    print(f"  精确度: {overall_relation['precision']:.4f}")
    print(f"  召回率: {overall_relation['recall']:.4f}")
    print(f"  F1分数: {overall_relation['f1']:.4f}")

    # 打印每种实体类型的总体指标
    print("\n每种实体类型的总体指标:")
    for entity_type, stats in overall_entity_type_stats.items():
        print(f"  {entity_type}: TP={stats['tp']}, FP={stats['fp']}, FN={stats['fn']}, "
              f"精确度={stats['precision']:.4f}, 召回率={stats['recall']:.4f}, F1={stats['f1']:.4f}")

    # 打印每种关系类型的总体指标
    print("\n每种关系类型的总体指标:")
    for relation_type, stats in overall_relation_type_stats.items():
        print(f"  {relation_type}: TP={stats['tp']}, FP={stats['fp']}, FN={stats['fn']}, "
              f"精确度={stats['precision']:.4f}, 召回率={stats['recall']:.4f}, F1={stats['f1']:.4f}")

    # 特别检查中成药的指标，如果不为1则打印
    print("\n" + "=" * 60)
    print("中成药指标检查")
    print("=" * 60)

    if "中成药" in overall_entity_type_stats:
        cm_stats = overall_entity_type_stats["中成药"]
        precision_cm = cm_stats['precision']
        recall_cm = cm_stats['recall']
        f1_cm = cm_stats['f1']

        print(f"中成药实体识别指标:")
        print(f"  精确度: {precision_cm:.4f}")
        print(f"  召回率: {recall_cm:.4f}")
        print(f"  F1分数: {f1_cm:.4f}")

        # 检查哪些指标不为1
        not_perfect_metrics = []
        if precision_cm != 1.0:
            not_perfect_metrics.append(f"精确度({precision_cm:.4f})")
        if recall_cm != 1.0:
            not_perfect_metrics.append(f"召回率({recall_cm:.4f})")
        if f1_cm != 1.0:
            not_perfect_metrics.append(f"F1分数({f1_cm:.4f})")

        if not_perfect_metrics:
            print(f"注意: 中成药实体的以下指标不为1: {', '.join(not_perfect_metrics)}")
        else:
            print("中成药实体的所有指标均为1，表现完美！")

        # 打印具体的中成药未匹配实体
        if overall_cm_fn_entities:
            print(f"\n未正确识别的中成药实体 (FN - 标准中有但预测中没有):")
            for entity in overall_cm_fn_entities:
                print(f"  {entity}")

        if overall_cm_fp_entities:
            print(f"\n错误识别的中成药实体 (FP - 预测中有但标准中没有):")
            for entity in overall_cm_fp_entities:
                print(f"  {entity}")

        if not overall_cm_fn_entities and not overall_cm_fp_entities:
            print("\n所有中成药实体都被正确识别！")

    else:
        print("未找到中成药实体的统计信息")

    # 整理所有结果
    all_results = {
        'file_results': file_results,
        'overall_triplet': overall_triplet,
        'overall_entity': overall_entity,
        'overall_relation': overall_relation,
        'overall_entity_type_stats': overall_entity_type_stats,
        'overall_relation_type_stats': overall_relation_type_stats
    }

    # 保存到Excel文件
    if save_excel:
        if output_file is None:
            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            output_file = f"评估结果_{timestamp}.xlsx"

        save_results_to_excel(all_results, output_file)

    return all_results


if __name__ == "__main__":
    # 设置文件夹路径
    gold_folder = "verification"  # 人工标注文件夹
    pred_folder = "doubao"  # 模型输出文件夹

    # 确保文件夹存在
    if not os.path.exists(gold_folder):
        print(f"错误: 文件夹 {gold_folder} 不存在")
        exit(1)

    if not os.path.exists(pred_folder):
        print(f"错误: 文件夹 {pred_folder} 不存在")
        exit(1)

    print("=" * 60)
    print("DeepSeek-R1模型性能评估")
    print("=" * 60)

    # 处理文件夹并保存结果
    results = process_folders(gold_folder, pred_folder, save_excel=True)
